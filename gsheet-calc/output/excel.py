"""Excel workbook output generation."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.issue import ReviewIssue
from models.project import Project
from models.result import CalcResult
from models.scenario import ScenarioResult
from models.site import DataSource, Site
from output.templates.workbook_template import (
    ADVISORY_FILL_HEX,
    DETERMINISTIC_FILL_HEX,
    HEADER_FILL_HEX,
    HEADER_FONT_COLOR,
    MANUAL_REVIEW_FILL_HEX,
    get_footer_with_date,
)

_header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
_header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
_advisory_fill = PatternFill(start_color=ADVISORY_FILL_HEX, end_color=ADVISORY_FILL_HEX, fill_type="solid")
_review_fill = PatternFill(start_color=MANUAL_REVIEW_FILL_HEX, end_color=MANUAL_REVIEW_FILL_HEX, fill_type="solid")
_det_fill = PatternFill(start_color=DETERMINISTIC_FILL_HEX, end_color=DETERMINISTIC_FILL_HEX, fill_type="solid")


def _add_header_row(ws, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _header_font
        cell.fill = _header_fill
        cell.alignment = Alignment(horizontal="center")


def _add_footer(ws, row: int) -> None:
    """Add the standard disclaimer footer."""
    ws.cell(row=row + 2, column=1, value=get_footer_with_date()).font = Font(italic=True, size=9)


def _auto_width(ws) -> None:
    """Auto-fit column widths (approximate)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def generate_workbook(
    site: Site,
    project: Project,
    area_results: list[CalcResult],
    density_results: list[CalcResult],
    far_results: list[CalcResult],
    height_results: list[CalcResult],
    parking_results: list[CalcResult],
    open_space_results: list[CalcResult],
    loading_results: list[CalcResult],
    scenarios: list[ScenarioResult],
    issues: list[ReviewIssue],
    output_path: Path,
) -> Path:
    """Generate the full Excel workbook."""
    wb = Workbook()

    # --- Tab 1: Project Summary ---
    ws = wb.active
    ws.title = "Project Summary"
    _add_header_row(ws, ["Field", "Value"])
    summary_rows = [
        ("Project Name", project.project_name),
        ("Project Number", project.project_number),
        ("Address", site.address),
        ("APN", site.apn),
        ("Zone", site.zone),
        ("Height District", site.height_district),
        ("Total Units", project.total_units),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Issues (total)", len(issues)),
        ("Blocking Issues", sum(1 for i in issues if i.blocking)),
    ]
    for row, (field, val) in enumerate(summary_rows, 2):
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=str(val) if val is not None else "")
    _add_footer(ws, len(summary_rows) + 2)
    _auto_width(ws)

    # --- Tab 2: Site Data ---
    ws = wb.create_sheet("Site Data")
    _add_header_row(ws, ["Field", "Value", "Confidence"])
    site_dict = site.model_dump(exclude={"data_sources", "raw_source_files", "parcel_geometry"})
    for row, (k, v) in enumerate(site_dict.items(), 2):
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v) if v is not None else "")
    _add_footer(ws, len(site_dict) + 2)
    _auto_width(ws)

    # --- Tab 3: Data Sources ---
    ws = wb.create_sheet("Data Sources")
    _add_header_row(ws, ["Field", "Source", "URL", "Raw Ref", "Pull Date", "Confidence", "Notes"])
    for row, ds in enumerate(site.data_sources, 2):
        ws.cell(row=row, column=1, value=ds.field)
        ws.cell(row=row, column=2, value=ds.source)
        ws.cell(row=row, column=3, value=ds.source_url or "")
        ws.cell(row=row, column=4, value=ds.raw_reference or "")
        ws.cell(row=row, column=5, value=ds.pull_date or "")
        ws.cell(row=row, column=6, value=ds.confidence)
        ws.cell(row=row, column=7, value=ds.notes or "")
    _add_footer(ws, len(site.data_sources) + 2)
    _auto_width(ws)

    # --- Tab 4: Assumptions ---
    ws = wb.create_sheet("Assumptions")
    _add_header_row(ws, ["Field", "Value", "Source"])
    proj_dict = project.model_dump(exclude={"unit_mix", "occupancy_areas", "frontage_segments", "affordability"})
    for row, (k, v) in enumerate(proj_dict.items(), 2):
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v) if v is not None else "")
        ws.cell(row=row, column=3, value="user_input")
    _add_footer(ws, len(proj_dict) + 2)
    _auto_width(ws)

    # --- Tab 5: Issue Register ---
    ws = wb.create_sheet("Issue Register")
    _add_header_row(ws, ["ID", "Category", "Severity", "Status", "Blocking", "Title", "Description", "Review Role"])
    for row, issue in enumerate(issues, 2):
        ws.cell(row=row, column=1, value=issue.id)
        ws.cell(row=row, column=2, value=issue.category)
        ws.cell(row=row, column=3, value=issue.severity)
        ws.cell(row=row, column=4, value=issue.status)
        ws.cell(row=row, column=5, value="YES" if issue.blocking else "")
        ws.cell(row=row, column=6, value=issue.title)
        ws.cell(row=row, column=7, value=issue.description)
        ws.cell(row=row, column=8, value=issue.suggested_review_role)
        if issue.blocking:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = _review_fill
    _add_footer(ws, len(issues) + 2)
    _auto_width(ws)

    # --- Tab 6: Area Calculations ---
    _write_calc_tab(wb, "Area Calculations", area_results)

    # --- Tab 7: Base Density + FAR ---
    _write_calc_tab(wb, "Density + FAR", density_results + far_results)

    # --- Tab 8: Parking Summary ---
    _write_calc_tab(wb, "Parking Summary", parking_results)

    # --- Tab 9: Open Space + Loading ---
    _write_calc_tab(wb, "Open Space + Loading", open_space_results + loading_results)

    # --- Tab 10: Advisory Screens ---
    ws = wb.create_sheet("Advisory Screens")
    _add_header_row(ws, ["Scenario", "Status", "Determinism", "Summary", "Missing Inputs", "Unresolved"])
    for row, sc in enumerate(scenarios, 2):
        ws.cell(row=row, column=1, value=sc.name)
        ws.cell(row=row, column=2, value=sc.status)
        ws.cell(row=row, column=3, value=sc.determinism)
        ws.cell(row=row, column=4, value=sc.summary)
        ws.cell(row=row, column=5, value="; ".join(sc.missing_inputs))
        ws.cell(row=row, column=6, value="; ".join(sc.unresolved))
        if sc.determinism == "advisory":
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = _advisory_fill
    _add_footer(ws, len(scenarios) + 2)
    _auto_width(ws)

    # --- Tab 11: Code References ---
    ws = wb.create_sheet("Code References")
    _add_header_row(ws, ["Calc Name", "Code Section", "Authority ID", "Code Cycle"])
    all_calcs = area_results + density_results + far_results + height_results + parking_results + open_space_results + loading_results
    for row, cr in enumerate(all_calcs, 2):
        ws.cell(row=row, column=1, value=cr.name)
        ws.cell(row=row, column=2, value=cr.code_section or "")
        ws.cell(row=row, column=3, value=cr.authority_id or "")
        ws.cell(row=row, column=4, value=cr.code_cycle)
    _add_footer(ws, len(all_calcs) + 2)
    _auto_width(ws)

    # --- Tab 12: Raw Source Log ---
    ws = wb.create_sheet("Raw Source Log")
    _add_header_row(ws, ["File", "Pull Timestamp"])
    for row, f in enumerate(site.raw_source_files, 2):
        ws.cell(row=row, column=1, value=f)
    ws.cell(row=2, column=2, value=site.pull_timestamp or "")
    _add_footer(ws, max(3, len(site.raw_source_files) + 2))
    _auto_width(ws)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


def _write_calc_tab(wb: Workbook, title: str, calcs: list[CalcResult]) -> None:
    """Write a calculation results tab."""
    ws = wb.create_sheet(title)
    _add_header_row(ws, [
        "Name", "Value", "Unit", "Formula", "Determinism", "Confidence",
        "Code Section", "Authority ID", "Steps", "Notes",
    ])
    for row, cr in enumerate(calcs, 2):
        ws.cell(row=row, column=1, value=cr.name)
        ws.cell(row=row, column=2, value=str(cr.value) if cr.value is not None else "N/A")
        ws.cell(row=row, column=3, value=cr.unit)
        ws.cell(row=row, column=4, value=cr.formula)
        ws.cell(row=row, column=5, value=cr.determinism)
        ws.cell(row=row, column=6, value=cr.confidence)
        ws.cell(row=row, column=7, value=cr.code_section or "")
        ws.cell(row=row, column=8, value=cr.authority_id or "")
        ws.cell(row=row, column=9, value="; ".join(cr.intermediate_steps))
        ws.cell(row=row, column=10, value="; ".join(cr.review_notes))

        if cr.determinism == "advisory":
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = _advisory_fill
        elif cr.confidence == "low":
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = _review_fill

    _add_footer(ws, len(calcs) + 2)
    _auto_width(ws)
